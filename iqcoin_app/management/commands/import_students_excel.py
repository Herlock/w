import openpyxl
from django.core.management.base import BaseCommand
from django.contrib.auth.models import User
from iqcoin_app.models import Student, UserProfile


class Command(BaseCommand):
    help = 'Import students from Excel file "Ученики Айкьюшки.xlsx"'
    
    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='Ученики Айкьюшки.xlsx',
            help='Path to the Excel file (default: Ученики Айкьюшки.xlsx)',
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default='Лист1',
            help='Sheet name to import from (default: Лист1)',
        )
    
    def handle(self, *args, **options):
        file_path = options['file']
        sheet_name = options['sheet']
        
        try:
            # Load the workbook and select the sheet
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook[sheet_name]
            
            # Get all teacher full names and map them to User objects
            teacher_mapping = {}
            teachers = User.objects.filter(userprofile__role='teacher')
            for teacher in teachers:
                profile = teacher.userprofile
                if profile.full_name:
                    teacher_mapping[profile.full_name.strip()] = teacher
                # Also map by username as fallback
                teacher_mapping[teacher.username] = teacher
            
            self.stdout.write(f"Found {len(teacher_mapping)} teachers in the database")
            
            # Parse the header row
            header_row = next(worksheet.iter_rows(values_only=True))
            headers = [str(cell).strip() if cell else '' for cell in header_row]
            
            # Expected headers mapping
            header_mapping = {
                'student_name': ['student_name', 'имя_ученика', 'student_name\xa0'],
                'teacher_full_name': ['teacher_full_name', 'учитель', 'teacher_full_name '],
                'phone_number': ['phone_number', 'номер_телефона', 'phone_number '],
                'is_active': ['is_active', 'активен', 'is_active\xa0'],
                'is_hidden': ['is_hidden', 'скрыт', 'is_hidden\xa0'],
            }
            
            # Map column indices
            column_indices = {}
            for i, header in enumerate(headers):
                for field, possible_headers in header_mapping.items():
                    if header in possible_headers:
                        column_indices[field] = i
                        break
            
            # Check if required columns are present
            if 'student_name' not in column_indices:
                self.stdout.write(
                    self.style.ERROR("Required column 'student_name' not found in the Excel file")
                )
                return
            
            if 'teacher_full_name' not in column_indices:
                self.stdout.write(
                    self.style.ERROR("Required column 'teacher_full_name' not found in the Excel file")
                )
                return
            
            # Process each row
            imported_count = 0
            skipped_count = 0
            error_count = 0
            
            # Skip header row and process data rows
            for row_num, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                # Skip header row
                if row_num == 1:
                    continue
                
                # Skip empty rows
                if not any(cell is not None for cell in row):
                    continue
                
                try:
                    # Extract values from row
                    student_name = self._get_cell_value(row, column_indices, 'student_name')
                    teacher_name = self._get_cell_value(row, column_indices, 'teacher_full_name')
                    phone_number = self._get_cell_value(row, column_indices, 'phone_number')
                    is_active_str = self._get_cell_value(row, column_indices, 'is_active')
                    is_hidden_str = self._get_cell_value(row, column_indices, 'is_hidden')
                    
                    # Skip if student name is empty
                    if not student_name:
                        skipped_count += 1
                        continue
                    
                    # Process boolean values
                    is_active = self._parse_boolean(is_active_str, default=True)
                    is_hidden = self._parse_boolean(is_hidden_str, default=False)
                    
                    # Find teacher
                    teacher = None
                    if teacher_name:
                        teacher_name_clean = teacher_name.strip()
                        if teacher_name_clean in teacher_mapping:
                            teacher = teacher_mapping[teacher_name_clean]
                        else:
                            # Try to find teacher by username
                            try:
                                teacher = User.objects.get(username=teacher_name_clean)
                            except User.DoesNotExist:
                                self.stdout.write(
                                    self.style.WARNING(
                                        f"Teacher '{teacher_name_clean}' not found for student '{student_name}' on row {row_num}. Skipping."
                                    )
                                )
                                error_count += 1
                                continue
                    
                    # If no teacher found, skip this student
                    if not teacher:
                        self.stdout.write(
                            self.style.WARNING(
                                f"No teacher found for student '{student_name}' on row {row_num}. Skipping."
                            )
                        )
                        error_count += 1
                        continue
                    
                    # Create or update student
                    student, created = Student.objects.get_or_create(
                        name=student_name.strip(),
                        teacher=teacher,
                        defaults={
                            'phone_number': phone_number.strip() if phone_number else None,
                            'is_active': is_active,
                            'is_hidden': is_hidden,
                            'balance': 0,
                        }
                    )
                    
                    if created:
                        imported_count += 1
                        self.stdout.write(
                            self.style.SUCCESS(
                                f"Imported student '{student_name}' assigned to teacher '{teacher_name}'"
                            )
                        )
                    else:
                        # Update existing student
                        student.phone_number = phone_number.strip() if phone_number else None
                        student.is_active = is_active
                        student.is_hidden = is_hidden
                        student.save()
                        self.stdout.write(
                            self.style.SUCCESS(
                                f"Updated student '{student_name}'"
                            )
                        )
                        imported_count += 1
                        
                except Exception as e:
                    self.stdout.write(
                        self.style.ERROR(
                            f"Error processing row {row_num}: {str(e)}"
                        )
                    )
                    error_count += 1
            
            # Summary
            self.stdout.write(
                self.style.SUCCESS(
                    f"Import completed. Imported/Updated: {imported_count}, Skipped: {skipped_count}, Errors: {error_count}"
                )
            )
            
        except FileNotFoundError:
            self.stdout.write(
                self.style.ERROR(f"File '{file_path}' not found")
            )
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f"Error importing students: {str(e)}")
            )
    
    def _get_cell_value(self, row, column_indices, field_name):
        """Helper method to safely get cell value"""
        if field_name in column_indices and column_indices[field_name] < len(row):
            value = row[column_indices[field_name]]
            return str(value) if value is not None else None
        return None
    
    def _parse_boolean(self, value, default=False):
        """Helper method to parse boolean values from various formats"""
        if value is None:
            return default
        
        value_str = str(value).strip().lower()
        if value_str in ['yes', 'true', '1', 'да', 'active']:
            return True
        elif value_str in ['no', 'false', '0', 'нет', 'inactive']:
            return False
        else:
            return default